from collections import defaultdict
from typing import Dict, DefaultDict, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


SENDER_COLUMN = "J"
RECIPIENT_COLUMN = "M"
POSTCARD_ID_COLUMN = "A"
HEADER_ROW = 1
ILLEGIBLE_TAG = "[нрзб]"
MISSING_TAG = "[отсутствует]"

def compilate_data(
    path: str,
    sheet_name: str | None = None,
    postcard_id_col: str = POSTCARD_ID_COLUMN,
    sender_col: str = SENDER_COLUMN,
    recipient_col: str = RECIPIENT_COLUMN,
) -> Tuple[Dict[int, str], Dict[int, str]]:
    """Читает данные из .xlsx-файла, составляет пары данных по соответствию.

    Читает столбцы с номерами открыток, получателями и отправителями из .xlsx-файла. Составляет пары "номер открытки" + "получатель", "номер открытки" + "отправитель", сохраняет их в два отдельных словаря.

    Args:
        path (str): путь к .xlsx-файлу.
        sheet_name (str | None): имя используемого листа, по умолчанию активный.
        postcard_id_col (str): буква столбца, содержащего номера открыток.
        sender_col (str): буква столбца, содержащего имена отправителей.
        recipient_col (str): буква столбца, содержащего имена получателей.

    Returns:
        Tuple[Dict[int, str], Dict[int, str]]: словарь с парами "номер открытки" + "отправитель", словарь с парами "номер открытки" + "получатель".
    """
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.active

    sender_id = column_index_from_string(sender_col)
    recipient_id = column_index_from_string(recipient_col)
    postcard_id = column_index_from_string(postcard_id_col)

    senders_raw: Dict[int, str] = {}
    recipients_raw: Dict[int, str] = {}

    for row in sheet.iter_rows(min_row=HEADER_ROW + 1):
        sender_val = row[sender_id - 1].value
        recipient_val = row[recipient_id - 1].value
        postcard_id_val = row[postcard_id - 1].value
        if postcard_id_val is not None:
            postcard_id_val = int(postcard_id_val)
            senders_raw[postcard_id_val] = "" if sender_val is None else str(sender_val).strip()
            recipients_raw[postcard_id_val] = "" if recipient_val is None else str(recipient_val).strip()

    return senders_raw, recipients_raw

def compilate_grouped_counts(
        anchor_ids: Dict[int, str],
        linked_ids: Dict[int, str],
) -> Dict[str, Dict[str, int]]:
    """Попарно группирует опорные и связанные значения по соответствиям, считает количество соответствий.

    Игнорирует тег [нрзб] и тег [отсутствует] в качестве опорного значения.

    Args:
        anchor_ids (Dict[int, str]): опорные значения.
        linked_ids (Dict[int, str]): связанные значения.

    Returns:
        Dict[str, Dict[str, int]]: {опорное значение: {связанное значение: количество соответствий}}
    """
    grouped_count: DefaultDict[str, DefaultDict[str, int]] = defaultdict(lambda: defaultdict(int))

    for postcard_id, anchor in anchor_ids.items():
        if anchor not in (ILLEGIBLE_TAG, MISSING_TAG):
            grouped_count[anchor][linked_ids[postcard_id]] += 1

    return {anchor: dict(linked) for anchor, linked in grouped_count.items()}

def count_unsufficient_tags(
        ids: Dict[int, str] = None,
) -> Tuple[int, int]:
    """Считает кол-во тегов [отсутствует] и [нрзб].

    Args:
        ids (Dict[int, str]): словарь соответствий "номер открытки" + "значение" ("получатель" или "отправитель").

    Returns:
        Tuple[int, int]: кол-во тегов [отсутствует], кол-во тегов [нрзб].
    """
    missing_tag_count = 0
    illegible_tag_count = 0

    for entry in ids.values():
        if entry == MISSING_TAG:
            missing_tag_count += 1
        elif entry == ILLEGIBLE_TAG:
            illegible_tag_count += 1

    return missing_tag_count, illegible_tag_count

def table_compilate(
        wb: Workbook,
        grouped_counts: Dict[str, Dict[str, int]],
        missed_tag_value: int = 0,
        illegible_tag_value: int = 0,
        sheet_title: str = "Соответствия",
        anchor_data_header: str = "Основа",
        linked_data_header: str = "Соответствия",
        pairs_quantity_header: str = "Кол-во соответствий"
) -> None:
    """Cоставляет таблицу соответствий.

    Первый столбец - опорные данные, второй столбец - найденные соответствия, третий столбец - кол-во найденных соответствий.

    Args:
        wb (Workbook): используемая книга.
        grouped_counts (Dict[str, Dict[str, int]]): опорные данные, установленные соответствия и кол-во установленных соответствий.
        missed_tag_value (int): кол-во тегов [отсутствует] в опорных данных.
        illegible_tag_value (int): кол-во тегов [нрзб] в опорных данных.
        sheet_title (str): название листа.
        anchor_data_header (str): заголовок столбца с опорными данными.
        linked_data_header (str): заголовок столбца с найденными соответствиями.
        pairs_quantity_header (str): заголовок столбца с кол-вом найденных соответствий.

    Returns:
        None
    """
    current_line = 2
    
    result: Worksheet = wb.create_sheet(title=sheet_title)
    result.column_dimensions["A"].width = 40
    result.column_dimensions["B"].width = 40
    result.column_dimensions["E"].width = 30
    result.column_dimensions["F"].width = 30

    result["A1"] = anchor_data_header
    result["B1"] = linked_data_header
    result["C1"] = pairs_quantity_header
    result["E1"] = f"{anchor_data_header}: [нрзб]"
    result["E2"] = illegible_tag_value
    result["F1"] = f"{anchor_data_header}: [отсутствует]"
    result["F2"] = missed_tag_value

    for anchor, linked in grouped_counts.items():
        result[f"A{current_line}"] = anchor
        if len(linked) > 1:
            result.merge_cells(f"A{current_line}:A{current_line + len(linked) - 1}")
        for name, amount in linked.items():
            result[f"B{current_line}"] = name
            result[f"C{current_line}"] = amount
            current_line += 1

    return